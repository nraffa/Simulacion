from openpyxl import Workbook
'''
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename = "hello_world.xlsx")

'''
from openpyxl import load_workbook
import json

workbook = load_workbook(filename = "reviews-sample.xlsx")
sheet = workbook.active

products = {}

'''
for row in sheet.iter_rows(min_row = 2, min_col = 4, max_col = 7, values_only = True ):
    
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3],
    }
    products[product_id] = product 

print(json.dumps(products, indent = 2))
'''
headers = []

for value in sheet.iter_rows(min_row = 1, max_row = 1, min_col =4, max_col= 7 , values_only = True):
    headers = value
print(headers)

products = {}

i=0
for row in sheet.iter_rows(min_row = 2, min_col = 4, max_col = 7, values_only = True ):
    product_id = row[0]
    for i in range(len(headers)):
        products[headers[i]] = row[i]




    print(json.dumps(products, indent = 2))

print (type(sheet['A15'].value))