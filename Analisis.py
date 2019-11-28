#from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook as wb
import json
from datetime import datetime

name = str(input("Escribir el nombre del archivo exacto(incluyendo el formato):"))

exfile = load_workbook(filename = name)
sheet = exfile[str(input("Ingrese nombre hoja:"))]

headers = []
a = int(input ("Ingrese cantidad de columnas:"))
b = int(input ("Ingrese cantidad de filas:"))
for value in sheet.iter_rows(min_row = 1, max_row = 1, min_col =1, max_col= a , values_only = True):
    headers = value
print (headers)

i=0
datas = {}
for row in sheet.iter_rows(min_row = 2, max_row = b, min_col = 1, max_col = a ,  values_only = True):
    
    data_id = row[0] #supongo que id esta en columna A
    for i in range(len(headers)):
        if type(row[i]) == datetime:
            data_date = row[i]
            corrected_date = datetime.str
        datas[headers[i]] = row[i]
        

print(json.dumps(datas, indent = 2))